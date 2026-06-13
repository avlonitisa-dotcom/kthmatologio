"""
FastAPI routes - all API endpoints.
"""
import asyncio
import csv
import io
import json
import logging
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, StreamingResponse

from config import Config
from database import db
from api.schemas import (
    AddKaekRequest, ImportKaeksRequest, GenerateTestRequest,
    ProcessRequest, AreaSearchRequest, FilterRequest,
    ConfigUpdateRequest,
)
from automation.kaek_workflow import generate_test_kaeks, sanitize_kaek, process_batch
from area_search.map_client import (
    discover_parcels, parse_kaek_list_from_text, search_municipalities,
)
from parsing.pdf_parser import parse_pdf_file

logger = logging.getLogger(__name__)
router = APIRouter()

# ── WebSocket connection manager ───────────────────────────────────────────────

class ConnectionManager:
    def __init__(self):
        self._clients: list[WebSocket] = []
        self._lock = asyncio.Lock()

    async def connect(self, ws: WebSocket):
        await ws.accept()
        async with self._lock:
            self._clients.append(ws)

    async def disconnect(self, ws: WebSocket):
        async with self._lock:
            self._clients.remove(ws)

    async def broadcast(self, data: dict):
        msg = json.dumps(data)
        dead = []
        for ws in self._clients:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            await self.disconnect(ws)


manager = ConnectionManager()

# Active processing task + stop event
_active_task: Optional[asyncio.Task] = None
_stop_event = asyncio.Event()


@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()  # Keep alive
    except WebSocketDisconnect:
        await manager.disconnect(websocket)


# ── Status & Config ────────────────────────────────────────────────────────────

@router.get("/api/status")
async def get_status():
    missing = Config.validate()
    return {
        "configured": len(missing) == 0,
        "missing_env": missing,
        "headless": Config.HEADLESS,
        "min_delay": Config.MIN_DELAY,
        "max_delay": Config.MAX_DELAY,
        "max_retries": Config.MAX_RETRIES,
        "llm_enabled": Config.LLM_ENABLED,
        "processing": _active_task is not None and not _active_task.done(),
    }


@router.post("/api/config")
async def update_config(req: ConfigUpdateRequest):
    if req.min_delay is not None:
        Config.MIN_DELAY = req.min_delay
    if req.max_delay is not None:
        Config.MAX_DELAY = req.max_delay
    if req.headless is not None:
        Config.HEADLESS = req.headless
    if req.max_retries is not None:
        Config.MAX_RETRIES = req.max_retries
    return {"ok": True}


# ── KAEK Management ────────────────────────────────────────────────────────────

@router.post("/api/kaek/add")
async def add_kaek(req: AddKaekRequest):
    kaek = sanitize_kaek(req.kaek)
    if not kaek:
        raise HTTPException(400, "Invalid KAEK format")
    db.upsert_kaek(kaek)
    db.add_log(kaek, "add", "info", "Added manually")
    return {"ok": True, "kaek": kaek}


@router.post("/api/kaek/import")
async def import_kaeks(req: ImportKaeksRequest):
    kaeks = []
    if req.kaeks:
        kaeks.extend([sanitize_kaek(k) for k in req.kaeks if k.strip()])
    if req.text:
        kaeks.extend(parse_kaek_list_from_text(req.text))

    kaeks = list(dict.fromkeys(k for k in kaeks if k))  # dedup, preserve order
    if not kaeks:
        raise HTTPException(400, "No valid KAEKs found")

    count = db.upsert_kaek_batch(kaeks)
    db.add_log(None, "import", "info", f"Imported {count} new KAEKs")
    return {"ok": True, "imported": count, "total": len(kaeks)}


@router.post("/api/kaek/generate-test")
async def generate_test(req: GenerateTestRequest):
    kaeks = generate_test_kaeks(req.prefix, req.count)
    count = db.upsert_kaek_batch(kaeks)
    db.add_log(None, "generate_test", "info", f"Generated {count} test KAEKs (UI testing only)")
    return {"ok": True, "kaeks": kaeks, "added": count}


@router.get("/api/kaek/list")
async def list_kaeks(page: int = 1, per_page: int = 100):
    jobs, total = db.get_all_kaeks(page, per_page)
    return {"jobs": jobs, "total": total, "page": page, "per_page": per_page}


@router.delete("/api/kaek/{kaek:path}")
async def delete_kaek(kaek: str):
    db.delete_kaek(kaek)
    return {"ok": True}


@router.post("/api/kaek/{kaek:path}/retry")
async def retry_kaek(kaek: str):
    db.reset_kaek_to_pending(kaek)
    return {"ok": True}


@router.post("/api/kaek/{kaek:path}/parse")
async def parse_kaek_pdfs(kaek: str):
    results = {}
    for pdf_type, path_field in [
        ("perigrafiki", "pdf_perigrafiki_path"),
        ("xoriki", "pdf_xoriki_path"),
    ]:
        jobs, _ = db.get_all_kaeks()
        job = next((j for j in jobs if j["kaek"] == kaek), None)
        if job and job.get(path_field):
            result = parse_pdf_file(kaek, pdf_type, Path(job[path_field]))
            results[pdf_type] = result
    return {"ok": True, "results": results}


# ── Processing Control ─────────────────────────────────────────────────────────

@router.post("/api/process/start")
async def start_processing(req: ProcessRequest):
    global _active_task, _stop_event

    if _active_task and not _active_task.done():
        raise HTTPException(409, "Processing already in progress")

    missing = Config.validate()
    if missing:
        raise HTTPException(400, f"Missing credentials: {missing}")

    if req.kaeks:
        kaeks = [sanitize_kaek(k) for k in req.kaeks if k.strip()]
    else:
        kaeks = [j["kaek"] for j in db.get_pending_kaeks()]

    if not kaeks:
        raise HTTPException(400, "No pending KAEKs to process")

    _stop_event = asyncio.Event()

    async def run_processing():
        from automation.browser_manager import start_browser, stop_browser, new_context
        try:
            await start_browser()
            ctx = await new_context()
            await process_batch(ctx, kaeks, manager.broadcast, _stop_event)
            await ctx.close()
        except Exception as exc:
            logger.error("Processing batch failed: %s", exc, exc_info=True)
            db.add_log(None, "batch", "failure", str(exc))
            await manager.broadcast({"type": "error", "message": str(exc)})
        finally:
            await stop_browser()
            await manager.broadcast({"type": "processing_done"})

    _active_task = asyncio.create_task(run_processing())
    return {"ok": True, "kaek_count": len(kaeks)}


@router.post("/api/process/stop")
async def stop_processing():
    global _stop_event
    _stop_event.set()
    return {"ok": True, "message": "Stop signal sent"}


# ── Dashboard ──────────────────────────────────────────────────────────────────

@router.get("/api/dashboard")
async def get_dashboard():
    stats = db.get_dashboard_stats()
    return stats


# ── Logs ───────────────────────────────────────────────────────────────────────

@router.get("/api/logs")
async def get_logs(kaek: Optional[str] = None, limit: int = 200, offset: int = 0):
    logs = db.get_logs(kaek=kaek, limit=limit, offset=offset)
    return {"logs": logs, "count": len(logs)}


# ── Municipality autocomplete ─────────────────────────────────────────────────

@router.get("/api/municipalities/search")
async def municipalities_search(q: str = "", limit: int = 10):
    """Autocomplete Greek municipality names via Nominatim."""
    if len(q) < 2:
        return {"results": []}
    results = await search_municipalities(q, limit=limit)
    return {"results": results}


# ── Area Search ────────────────────────────────────────────────────────────────

@router.post("/api/area-search/smart")
async def smart_search(req: AreaSearchRequest):
    """
    Full end-to-end smart search:
    map discovery → area filter → TEE download → PDF parse → results.
    """
    from area_search.smart_search import SmartSearchRequest, run_smart_search

    sr = SmartSearchRequest(
        area_name=req.area_name or "",
        map_url=req.map_url or "",
        target_sqm=req.min_area_sqm,   # Use min as target mid-point if only one bound
        min_sqm=req.min_area_sqm,
        max_sqm=req.max_area_sqm,
        tolerance_pct=10.0,
        has_building=req.has_building,
        has_burdens=req.has_burdens,
        kaek_pattern=req.kaek_pattern,
        manual_kaeks=req.kaek_list or [],
        download_pdfs=req.download_missing,
    )
    if req.kaek_text:
        from area_search.map_client import parse_kaek_list_from_text
        sr.manual_kaeks.extend(parse_kaek_list_from_text(req.kaek_text))

    _stop_event = asyncio.Event()
    results = await run_smart_search(sr, manager.broadcast, _stop_event)
    return {
        "ok": True,
        "total": len(results),
        "matching": sum(1 for r in results if r.matches_criteria),
        "results": [vars(r) for r in results],
    }


@router.post("/api/area-search/discover")
async def area_search_discover(req: AreaSearchRequest):
    """
    Discover KAEKs in a geographic area and apply area filter immediately.
    Building/burden filters are applied later (require PDF parsing).
    """
    from area_search.map_client import MapParcel, filter_by_area, filter_by_kaek_pattern

    manual_kaeks: list[str] = []
    if req.kaek_list:
        manual_kaeks.extend([sanitize_kaek(k) for k in req.kaek_list if k.strip()])
    if req.kaek_text:
        manual_kaeks.extend(parse_kaek_list_from_text(req.kaek_text))

    all_parcels: list[MapParcel] = [MapParcel(kaek=k) for k in manual_kaeks if k]

    # Discover from municipality name via ArcGIS REST (no browser needed)
    total_discovered = 0
    if req.area_name:
        try:
            ctx = None
            if Config.MAP_BROWSER_ENABLED:
                from automation.browser_manager import start_browser, new_context, stop_browser
                await start_browser()
                ctx = await new_context()
            parcels = await discover_parcels(municipality_name=req.area_name, ctx=ctx)
            total_discovered = len(parcels)
            all_parcels.extend(parcels)
            if ctx:
                await ctx.close()
                await stop_browser()
        except Exception as exc:
            logger.warning("Map discovery failed: %s", exc)
            db.add_log(None, "area_search", "warning", str(exc))

    # Apply area filter (we have area data from ArcGIS)
    filtered = filter_by_area(
        all_parcels,
        target_sqm=None,
        min_sqm=req.min_area_sqm,
        max_sqm=req.max_area_sqm,
    )

    # Apply KAEK pattern filter
    if req.kaek_pattern:
        filtered = filter_by_kaek_pattern(filtered, req.kaek_pattern)

    # Deduplicate
    seen: set[str] = set()
    unique: list[MapParcel] = []
    for p in filtered:
        if p.kaek not in seen:
            seen.add(p.kaek)
            unique.append(p)

    kaeks = [p.kaek for p in unique]

    if kaeks:
        count = db.upsert_kaek_batch(kaeks)
        db.add_log(None, "area_search", "info",
                   f"Discovered {total_discovered} parcels for '{req.area_name}', "
                   f"{len(kaeks)} after area filter, {count} new in DB")

    session_id = db.create_session({
        "area_name": req.area_name,
        "map_url": req.map_url,
        "min_area_sqm": req.min_area_sqm,
        "max_area_sqm": req.max_area_sqm,
        "has_building": req.has_building,
        "has_burdens": req.has_burdens,
        "kaek_pattern": req.kaek_pattern,
    })

    # Note: building/burden filters need PDF parsing — they apply after download
    area_filter_active = req.min_area_sqm or req.max_area_sqm
    return {
        "ok": True,
        "kaeks_discovered": len(kaeks),
        "total_in_area": total_discovered,
        "area_filter_applied": bool(area_filter_active),
        "note": (
            f"Φίλτρο εμβαδόν εφαρμόστηκε: {total_discovered} → {len(kaeks)} KAEK. "
            "Φίλτρα κτίσμα/βάρη εφαρμόζονται μετά τη λήψη PDF."
            if area_filter_active else
            "Βρέθηκαν όλα τα KAEK χωρίς φίλτρο εμβαδόν. "
            "Φίλτρα κτίσμα/βάρη εφαρμόζονται μετά τη λήψη PDF."
        ),
        "kaeks": kaeks[:100],  # Preview first 100
        "session_id": session_id,
    }


@router.post("/api/area-search/filter")
async def area_search_filter(req: FilterRequest):
    """Filter already-parsed results by criteria."""
    results = db.search_parse_results(
        min_area=req.min_area_sqm,
        max_area=req.max_area_sqm,
        has_building=req.has_building,
        has_burdens=req.has_burdens,
        only_completed=req.only_completed,
        kaek_pattern=req.kaek_pattern,
    )
    return {"results": results, "count": len(results)}


# ── PDF Viewing ────────────────────────────────────────────────────────────────

@router.get("/api/pdf/{kaek:path}/{pdf_type}")
async def serve_pdf(kaek: str, pdf_type: str):
    kaek_clean = kaek.replace("/", "_")
    if pdf_type == "perigrafiki":
        filename = f"{kaek_clean}_perigrafiki_vasi.pdf"
    elif pdf_type == "xoriki":
        filename = f"{kaek_clean}_xoriki_vasi.pdf"
    else:
        raise HTTPException(400, "Unknown pdf_type")

    path = Config.DOWNLOADS_DIR / kaek_clean / filename
    if not path.exists():
        raise HTTPException(404, "PDF not found")
    return FileResponse(str(path), media_type="application/pdf")


# ── Export ─────────────────────────────────────────────────────────────────────

@router.get("/api/export/csv")
async def export_csv():
    results = db.search_parse_results(only_completed=False)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "kaek", "pdf_type", "area_sqm", "area_stremma",
        "has_building", "has_burdens", "property_type",
        "burdens_detail", "confidence", "parse_method",
        "pdf_perigrafiki_path", "pdf_xoriki_path", "job_status",
    ])
    writer.writeheader()
    writer.writerows(results)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=kaek_results.csv"},
    )


@router.get("/api/export/excel")
async def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    results = db.search_parse_results(only_completed=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KAEK Results"

    headers = [
        "KAEK", "Εμβαδόν (τ.μ.)", "Εμβαδόν (στρ.)",
        "Κτίσμα", "Βάρη", "Τύπος", "Λεπτομέρειες Βαρών",
        "Εμπιστοσύνη", "Μέθοδος", "Κατάσταση",
        "PDF Περιγραφικής", "PDF Χωρικής",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.get("kaek"))
        ws.cell(row=row, column=2, value=r.get("area_sqm"))
        ws.cell(row=row, column=3, value=r.get("area_stremma"))
        ws.cell(row=row, column=4, value=r.get("has_building", "unknown"))
        ws.cell(row=row, column=5, value=r.get("has_burdens", "unknown"))
        ws.cell(row=row, column=6, value=r.get("property_type", ""))
        ws.cell(row=row, column=7, value=r.get("burdens_detail", ""))
        ws.cell(row=row, column=8, value=r.get("confidence"))
        ws.cell(row=row, column=9, value=r.get("parse_method", ""))
        ws.cell(row=row, column=10, value=r.get("job_status", ""))
        ws.cell(row=row, column=11, value=r.get("pdf_perigrafiki_path", ""))
        ws.cell(row=row, column=12, value=r.get("pdf_xoriki_path", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kaek_results.xlsx"},
    )
